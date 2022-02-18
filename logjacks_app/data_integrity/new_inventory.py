from copy import deepcopy
from codecs import iterdecode
from csv import reader, excel
from datetime import datetime, timezone
from openpyxl import load_workbook
from logjacks_app.data_integrity.config import *
from logjacks_app.timber import *
from logjacks_app.utils.formatting import back_date


def get_blank_row_json():
    return [i.__json__() for i in INVENTORY_ROW]


def imported_sheet_to_json(final_imported_data):
    master = []
    for row in final_imported_data:
        master.append([i.__json__() for i in row])
    return master


def read_imported_sheet(file):
    if file.filename.endswith('.csv'):
        no_error, data = _read_csv(file)
    else:
        wb = load_workbook(file, data_only=True)
        no_error, data = _read_excel(wb)
    return no_error, data


def data_to_timber(master):
    stands = []
    for stand_id, sub in master.items():
        stand = Stand(stand_id, sub['Acres'], back_date(sub['Date of Inventory']))
        hdr = _get_avg_hdr(sub['DBH'], sub['Total Height'])
        current_plot = 0
        if 'Pref Log Length' in sub:
            quick = True
        else:
            quick = False
        plot = None
        for idx, p_num in enumerate(sub['Plot Number']):
            if p_num != current_plot:
                if current_plot > 0:
                    stand.add_plot(plot)
                plot = Plot(sub['Plot Factor'][idx])
                current_plot = p_num
            plot.add_tree(construct_tree(sub, idx, hdr, plot.plot_factor, quick=quick))
        stand.add_plot(plot)
        stands.append(stand)
    return stands


def construct_tree(sub, idx, hdr, plot_factor, quick=True):
    spp = sub['Species'][idx]
    dbh = sub['DBH'][idx]
    height = sub['Total Height'][idx]
    if height == '':
        height = (dbh / 12) * hdr

    if quick:
        plog = sub['Pref Log Length'][idx]
        mlog = sub['Min Log Length'][idx]
        utdib = sub['Utility Log DIB'][idx]
        return TimberQuick(plot_factor, spp, dbh, height, plog, mlog, utdib)
    else:
        tree = TimberFull(plot_factor, spp, dbh, height)
        stem_height = sub['Stump Height'][idx]
        lnum = 1
        key = f'Log {lnum} Length'
        while key in sub:
            length = sub[key][idx]
            if length == '':
                break
            stem_height += sub[key][idx]
            tree.add_log(stem_height, length, sub[f'Log {lnum} Grade'][idx], sub[f'Log {lnum} Defect'][idx])
            stem_height += sub[f'Between Logs{lnum}'][idx] + 1

            lnum += 1
            key = f'Log {lnum} Length'
        return tree





def _get_avg_hdr(dbhs, heights):
    master = []
    for dbh, hgt in zip(dbhs, heights):
        if hgt != '':
            master.append(hgt / (dbh / 12))
    return sum(master) / len(master)




def _read_csv(file):
    convert = iterdecode(file.stream, 'utf-8')
    read = reader(convert, dialect=excel)
    headers = next(read)
    if headers:
        required = {
            key: {
                'idx': None,
                'possible_names': _possible_all_col_names(POSSIBLE_COL_NAMES_FROM_UPLOAD[key]['names'])
            }
            for key in POSSIBLE_COL_NAMES_FROM_UPLOAD if POSSIBLE_COL_NAMES_FROM_UPLOAD[key]['required']
        }
        has_required = True
        for key in required:
            pos_names = required[key]['possible_names']
            full_check = headers + pos_names
            filtered = list(filter(lambda x: True if full_check.count(x) == 2 and x in pos_names else False, full_check))
            if filtered:
                idx = headers.index(filtered[0])
                required[key]['idx'] = idx
            else:
                has_required = False

        if has_required:
            log_cols = _get_log_cols(headers)

            # Checking if there is at least one log, else try to find quick cruise cols
            if log_cols[1]['Log 1 Length']['idx'] is not None:
                full_cruise = True
                for log_num in log_cols:
                    required.update(log_cols[log_num])
            else:
                full_cruise = False
                del log_cols
                quick_cols = _get_quick_cols(headers)
                required.update(quick_cols)

            master = []
            for row in read:
                blank_row = deepcopy(INVENTORY_ROW)
                temp = []
                for elem in blank_row:
                    if elem.label in required:
                        idx = required[elem.label]['idx']
                        if elem.label.startswith('Log') and elem.label.endswith('Length'):
                            if idx is None:
                                break
                            elif row[required[elem.label]['idx']] == '':
                                break

                        if idx is None:
                            elem.val = required[elem.label]['default']
                        else:
                            elem.val = row[required[elem.label]['idx']]
                        temp.append(elem)
                master.append(temp)

            if master[-1][0].val == '':
                master.pop(-1)

            # Evening out the row lists with blank values to the max amount of logs per tree in the stand
            if full_cruise:
                max_row_len = max([len(row) for row in master])
                for row in master:
                    row_len = len(row)
                    if row_len < max_row_len:
                        blank_row = deepcopy(INVENTORY_ROW)
                        for elem in blank_row[row_len + 3: max_row_len + 3]:
                            if elem.label == 'Between Logs':
                                elem.val = ''
                            row.append(elem)

            return True, imported_sheet_to_json(master)

    return False, None


def _read_excel(workbook):
    # Checking for required columns first
    required = {
        key: {
            'idx': None,
            'possible_names': _possible_all_col_names(POSSIBLE_COL_NAMES_FROM_UPLOAD[key]['names'])
        }
        for key in POSSIBLE_COL_NAMES_FROM_UPLOAD if POSSIBLE_COL_NAMES_FROM_UPLOAD[key]['required']
    }
    for sheet in workbook.sheetnames:
        ws = workbook[sheet]
        headers = [i.upper() if i is not None else i for i in list(next(ws.values))]
        has_required = True
        for key in required:
            pos_names = required[key]['possible_names']
            full_check = headers + pos_names
            filtered = list(filter(lambda x: True if full_check.count(x) == 2 and x in pos_names else False, full_check))
            if filtered:
                idx = headers.index(filtered[0]) + 1
                required[key]['idx'] = idx
            else:
                has_required = False

        if has_required:
            log_cols = _get_log_cols(headers)

            # Checking if there is at least one log, else try to find quick cruise cols
            if log_cols[1]['Log 1 Length']['idx'] is not None:
                full_cruise = True
                for log_num in log_cols:
                    required.update(log_cols[log_num])
            else:
                full_cruise = False
                del log_cols
                quick_cols = _get_quick_cols(headers)
                required.update(quick_cols)

            # Getting the data from the worksheet with the acquired indexes
            master = []
            for i, row in enumerate(ws.iter_rows(), 2):
                blank_row = deepcopy(INVENTORY_ROW)
                temp = []
                for elem in blank_row:
                    if elem.label in required:
                        idx = required[elem.label]['idx']
                        if elem.label.startswith('Log') and elem.label.endswith('Length'):
                            if idx is None:
                                break
                            elif ws.cell(i, required[elem.label]['idx']).value is None:
                                break

                        if idx is None:
                            elem.val = required[elem.label]['default']
                        else:
                            val = ws.cell(i, required[elem.label]['idx']).value
                            elem.val = _convert_none(val)
                        temp.append(elem)
                master.append(temp)

            if master[-1][0].val == '':
                master.pop(-1)

            # Evening out the row lists with blank values to the max amount of logs per tree in the stand
            if full_cruise:
                max_row_len = max([len(row) for row in master])
                for row in master:
                    row_len = len(row)
                    if row_len < max_row_len:
                        blank_row = deepcopy(INVENTORY_ROW)
                        for elem in blank_row[row_len + 3: max_row_len + 3]:
                            if elem.label == 'Between Logs':
                                elem.val = ''
                            row.append(elem)

            # r1 = [x.label for x in master[0]]
            # print(len(r1), r1)
            # for row in master:
            #     rn = [x.val for x in row]
            #     print(len(rn), rn)
            # print('\n')

            return True, imported_sheet_to_json(master)
    else:
        return False, None


def _get_log_cols(headers):
    # Look for "Between Logs Cols" and getting their indexes in order
    pos_names = _possible_all_col_names(POSSIBLE_COL_NAMES_FROM_UPLOAD['Between Logs']['names'])
    copy_headers = [i for i in headers]
    btw_idxs = []
    for i in copy_headers:
        if i in pos_names:
            idx = copy_headers.index(i)
            btw_idxs.append(idx)
            copy_headers.pop(idx)

    # Next checking for log cols for full cruise, checks out to 10 logs
    log_cols = {
        0: {
            'Stump Height': {
                'idx': None,
                'possible_names': _possible_all_col_names(POSSIBLE_COL_NAMES_FROM_UPLOAD['Stump Height']['names']),
                'default': POSSIBLE_COL_NAMES_FROM_UPLOAD['Stump Height']['default']
            }
        }
    }

    for i in range(1, 11):
        log_cols[i] = {}
        for key in POSSIBLE_COL_NAMES_FROM_UPLOAD:
            if POSSIBLE_COL_NAMES_FROM_UPLOAD[key]['cruise'] == 'f' and key != 'Stump Height':
                idx = None
                if key == 'Between Logs':
                    if i < len(btw_idxs):
                        idx = btw_idxs[i - 1]
                log_cols[i][key.format(i) if '{}' in key else key] = {
                    'idx': idx,
                    'possible_names': _possible_all_col_names(POSSIBLE_COL_NAMES_FROM_UPLOAD[key]['names'], i if '{}' in key else None),
                    'default': POSSIBLE_COL_NAMES_FROM_UPLOAD[key]['default']
                }

    for log_num in log_cols:
        for key in log_cols[log_num]:
            if key != 'Between Logs':
                pos_names = log_cols[log_num][key]['possible_names']
                full_check = headers + pos_names
                filtered = list(filter(lambda x: True if full_check.count(x) == 2 and x in pos_names else False, full_check))
                if filtered:
                    idx = headers.index(filtered[0]) + 1
                    log_cols[log_num][key]['idx'] = idx

    # for num in log_cols:
    #     print(num)
    #     for key in log_cols[num]:
    #         print(f'\t{key}: {log_cols[num][key]}')
    #     print()

    return log_cols


def _get_quick_cols(headers):
    quick_cols = {
        key: {
            'idx': None,
            'possible_names': _possible_all_col_names(POSSIBLE_COL_NAMES_FROM_UPLOAD[key]['names']),
            'default': POSSIBLE_COL_NAMES_FROM_UPLOAD[key]['default']
        }
        for key in POSSIBLE_COL_NAMES_FROM_UPLOAD if POSSIBLE_COL_NAMES_FROM_UPLOAD[key]['cruise'] == 'q'
    }
    for key in quick_cols:
        pos_names = quick_cols[key]['possible_names']
        full_check = headers + pos_names
        filtered = list(filter(lambda x: True if full_check.count(x) == 2 and x in pos_names else False, full_check))
        if filtered:
            idx = headers.index(filtered[0]) + 1
            quick_cols[key]['idx'] = idx

    # for key in quick_cols:
    #     print(key)
    #     for sub in quick_cols[key]:
    #         print(f'{sub}: {quick_cols[key][sub]}')
    #     print()

    return quick_cols


def _convert_none(val):
    return val if val is not None else ''


def _possible_all_col_names(col_list, fill=None):
    if fill is None:
        master = [i for i in col_list]
        separators = ['', '_', '-', '.']
        for i in separators:
            add = [j.replace(' ', i) for j in col_list]
            master += add
        return list(set(master))
    else:
        master = [i.format(fill) for i in col_list]
        separators = ['', '_', '-', '.']
        for i in separators:
            add = [j.replace(' ', i).format(fill) for j in col_list]
            master += add
        return list(set(master))

