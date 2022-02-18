from datetime import datetime
from logjacks_app.data_integrity.config import POSSIBLE_COL_NAMES_FROM_UPLOAD
from logjacks_app.data_integrity.new_inventory import _read_excel
from openpyxl import Workbook, load_workbook
from random import randrange, choice
import os
import time
#65 90



if __name__ == '__main__':
    pass


    # path = 'C:/ZBEE490/Dev/other/logjacks/del_scratch'
    # r_path = 'C:/ZBEE490/Dev/other/logjacks/logjacks_app/static/blank_sheets'
    #
    #
    #
    # real = os.path.join(r_path, 'stand_data_full.xlsx')
    #
    # scratch = 'scratch_{}.xlsx'
    # q = 'stand_data_quick.xlsx'
    # f = 'stand_data_full.xlsx'
    # c = 'normal_csv_to_try_onestand.xlsx'
    #
    #
    # #for i in range(1, 26):
    # test = os.path.join(path, c)
    # wb = load_workbook(test)
    # read_excel(wb)



    # for i in range(1, 26):
    #     pos_cols = {key: value for key, value in POSSIBLE_COL_NAMES_FROM_UPLOAD.items()}
    #     pos_cols_list = list(pos_cols)
    #
    #     wb = Workbook()
    #     ws = wb.active
    #
    #     for j in range(1, 41):
    #         keep_cols = ['Log {} Length', 'Log {} Grade', 'Log {} Defect', 'Between Logs']
    #         pos_base = choice(pos_cols_list)
    #         if pos_base in keep_cols:
    #             pos_col = choice(pos_cols[pos_base]['names'])
    #         else:
    #             pos_col = choice(pos_cols[pos_cols_list.pop(pos_cols_list.index(pos_base))]['names'])
    #
    #         if '{}' in pos_col:
    #             pos_col = pos_col.format(randrange(1, 6))
    #
    #         if randrange(1, 5) == 3:
    #             pos_col = pos_col.replace(' ', choice(['', '_', '-', '.']))
    #
    #         rand_col = f'{chr(randrange(65, 90))}{chr(randrange(65, 90))}{chr(randrange(65, 90))}'
    #         choices = [pos_col if i < 9 else rand_col for i in range(10)]
    #         ws.cell(1, j, choice(choices))
    #
    #         for k in range(2, 12):
    #             ws.cell(k, j, randrange(1, 100))
    #
    #     wb.save(os.path.abspath(os.path.join(path, f'scratch_{i}.xlsx')))
    #
    #     print(f'{i}/25', sep='', end='\r')







    # x = '22.2'
    # y = '22.5g5'
    # z = '0.0034'
    # a = '2'
    # b = 'a!'
    # c = '2!!'
    #
    # func = lambda y: float(y) if ''.join(filter(lambda x: False if x == '.' else True, y)).isdigit() else False
    #
    # for i in [x, y, z, a, b, c]:
    #     print(f'{i} {func(i)} numeric')

    # for i in range(1, 32):
    #     print(f'{datetime(2022, 1, i): %Y-%m-%d}')
    #     print(f'{datetime(2022, 1, i): %Y/%m/%d}')
