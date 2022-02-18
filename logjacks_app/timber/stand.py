from os import (
    startfile,
    getcwd
)
from os.path import join
from io import BytesIO
from csv import (
    writer,
    excel
)
from openpyxl import (
    Workbook,
    load_workbook
)
from statistics import (
    mean,
    variance,
    stdev
)
from logjacks_app.timber.plot import Plot
from logjacks_app.timber.timber import (
    TimberQuick,
    TimberFull
)
from logjacks_app.timber.log import Log
# from treetopper.thin import (
#     ThinTPA,
#     ThinBA,
#     ThinRD
# )
# from treetopper._exceptions import TargetDensityError
# from treetopper.fvs import FVS
from logjacks_app.timber.constants import (
    math,
    ALL_SPECIES_NAMES,
    GRADE_SORT,
    LOG_LENGTHS,
    SORTED_HEADS
)
# from treetopper._utils import (
#     format_comma,
#     format_pct,
#     extension_check,
#     reorder_dict,
#     check_date,
#     add_logs_to_table_heads
# )
# from treetopper._import_from_sheets import import_from_sheet
# from treetopper._print_console import (
#     print_stand_species,
#     print_stand_logs,
#     print_stand_stats
# )
# from treetopper._print_pdf import PDF


class Stand(object):
    """The Stand Class represents a stand of timber that has had an inventory conducted on it. It should made up of plots (Plot Class)
       which contain trees (Timber Classes).

       The Stand class will run calculations and statistics of the current stand conditions and it will run calculations of the log
       merchantabilty for three metrics: logs per acre, log board feet per acre, and log cubic feet per acre, based on log grades,
       log length ranges and species.

       """

    def __init__(self, name: str, acres: float = None, date_inventory: str = None):
        self.name = name.upper()
        self.acres = acres
        self.date_inventory = date_inventory

        self.plots = []
        self.plot_count = 0

        self.tpa = 0
        self.ba_ac = 0
        self.qmd = 0
        self.rd_ac = 0
        self.bf_ac = 0
        self.cf_ac = 0
        self.avg_hgt = 0
        self.hdr = 0
        self.vbar = 0

        self.tpa_stats = {}
        self.ba_ac_stats = {}
        self.rd_ac_stats = {}
        self.bf_ac_stats = {}
        self.cf_ac_stats = {}

        self.species = {}
        self.species_gross = {}
        self.species_stats = {}

        self.logs = {}

        self.table_data = []

        self.summary_stand = []
        self.summary_logs = {}
        self.summary_stats = []

        self.metrics = ['tpa', 'ba_ac', 'rd_ac', 'bf_ac', 'cf_ac']
        self.attrs = ['_gross', '_stats', '']

    def __getitem__(self, attribute: str):
        return self.__dict__[attribute]

    # def console_report(self):
    #     """Prints a console-formatted string of the complete stand report"""
    #     print(self._compile_report_text())
    #
    # def get_pdf_report_bytes_io(self):
    #     pdf = self._compile_pdf_report()
    #     return BytesIO(pdf.output(dest='S').encode('latin-1'))
    #
    # def pdf_report(self, filename: str, directory: str = None, start_file_upon_creation: bool = False):
    #     """Exports a pdf of the complete stand report to a user specified directory or if directory is None,
    #     to the current working directory. Will open the created pdf report if start_file_upon_creation is True"""
    #     check = filename if filename[-4:] == '.pdf' else f'{filename}.pdf'
    #     if directory:
    #         file = join(directory, check)
    #     else:
    #         file = join(getcwd(), check)
    #
    #     pdf = self._compile_pdf_report()
    #     pdf.output(file, 'F')
    #     if start_file_upon_creation:
    #         startfile(file)

    def add_plot(self, plot: Plot):
        """Adds a plot to the stand's plots list and re-runs the calculations and statistics of the stand.
           plot argument needs to be the a Plot Class"""
        self.plots.append(plot)
        self.plot_count += 1

        for met in self.metrics:
            self._update_metrics(met)
        self.qmd = math.sqrt((self.ba_ac / self.tpa) / .005454)
        self.vbar = self.bf_ac / self.ba_ac

        self._update_species(plot)
        self._update_logs(plot)
        # self.table_data = self._update_table_data()
        #
        # self.summary_stand = self._update_summary_stand()
        # self.summary_logs = self._update_summary_logs()
        # self.summary_stats = self._update_summary_stats()

    # def import_sheet_quick(self, file_path: str):
    #     """Imports tree and plot data from a CSV or XLSX file for a quick cruise and adds that data to the stand"""
    #     plots = import_from_sheet(file_path, self.name, 'q')
    #     for plot_num in plots:
    #         plot = Plot()
    #         for tree in plots[plot_num]:
    #             plot.add_tree(TimberQuick(self.plot_factor, *tree))
    #         self.add_plot(plot)
    #
    # def import_sheet_full(self, file_path: str):
    #     """Imports tree and plot data from a CSV or XLSX file for a full cruise and adds that data to the stand"""
    #     plots = import_from_sheet(file_path, self.name, 'f')
    #     for plot_num in plots:
    #         plot = Plot()
    #         for tree_data in plots[plot_num]:
    #             args = tree_data[: -1]
    #             logs = tree_data[-1]
    #             tree = TimberFull(self.plot_factor, *args)
    #             for log in logs:
    #                 tree.add_log(*log)
    #             plot.add_tree(tree)
    #         self.add_plot(plot)
    #
    # def table_to_csv(self, filename: str, directory: str = None):
    #     """Creates or appends a CSV file with tree data from self.table_data"""
    #     check = extension_check(filename, '.csv')
    #     if directory:
    #         file = join(directory, check)
    #     else:
    #         file = join(getcwd(), check)
    #
    #     if isfile(file):
    #         allow = 'a'
    #         start = 1
    #     else:
    #         allow = 'w'
    #         start = 0
    #
    #     with open(file, allow, newline='') as csv_file:
    #         csv_write = writer(csv_file, dialect=excel)
    #         for i in self.table_data[start:]:
    #             csv_write.writerow(i)
    #
    # def table_to_excel(self, filename: str, directory: str = None):
    #     """Creates or appends an Excel file with tree data from self.table_data"""
    #     check = extension_check(filename, '.xlsx')
    #     if directory:
    #         file = join(directory, check)
    #     else:
    #         file = join(getcwd(), check)
    #
    #     if isfile(file):
    #         wb = load_workbook(file)
    #         ws = wb.active
    #         for i in self.table_data[1:]:
    #             ws.append(i)
    #         wb.save(file)
    #     else:
    #         wb = Workbook()
    #         ws = wb.active
    #         for i in self.table_data:
    #             ws.append(i)
    #         wb.save(file)

    def _update_metrics(self, metric: str):
        """Updates stand metrics based on the metric entered in the argument, used internally"""
        metric_list = [plot[metric] for plot in self.plots]
        stats = self._get_stats(metric_list)
        setattr(self, metric, stats['mean'])
        setattr(self, f'{metric}_stats', stats)

    def _update_species(self, plot):
        """Re-runs stand conditions calculations and statistics, used internally"""
        update_after = ['qmd', 'vbar', 'avg_hgt', 'hdr']
        if self.plot_count == 0:
            return
        else:
            for species in plot.species:
                if species not in self.species_gross:
                    for attr in self.attrs:
                        if attr == '_gross':
                            getattr(self, f'species{attr}')[species] = {met: [] for met in self.metrics}
                        else:
                            getattr(self, f'species{attr}')[species] = {met: 0 for met in self.metrics}
                for key in plot.species[species]:
                    if key not in update_after:
                        self.species_gross[species][key].append(plot.species[species][key])
            for species in self.species_gross:
                for key in self.species_gross[species]:
                    if key not in update_after:
                        data = self.species_gross[species][key]
                        if len(data) < self.plot_count:
                            data += ([0] * (self.plot_count - len(data)))
                        stats = self._get_stats(data)
                        self.species[species][key] = stats['mean']
                        self.species_stats[species][key] = stats
                self.species[species]['qmd'] = math.sqrt((self.species[species]['ba_ac'] / self.species[species]['tpa']) / 0.005454)
                self.species[species]['vbar'] = self.species[species]['bf_ac'] / self.species[species]['ba_ac']
                if species == 'totals_all':
                    self.species[species]['avg_hgt'] = mean([p.avg_hgt for p in self.plots])
                    self.species[species]['hdr'] = mean([p.hdr for p in self.plots])
                else:
                    trees = []
                    for p in self.plots:
                        for t in p.trees:
                            trees.append(t)
                    self.species[species]['avg_hgt'] = mean([t.total_height for t in trees if t.species == species])
                    self.species[species]['hdr'] = mean([t.hdr for t in trees if t.species == species])

    def _update_logs(self, plot):
        """Re-runs stand logs calculations, used internally"""
        if self.plot_count == 0:
            return
        else:
            subs = ['lpa', 'bf_ac', 'cf_ac']
            for species in plot.logs:
                if species not in self.logs:
                    self.logs[species] = {}
                for grade in plot.logs[species]:
                    if grade not in self.logs[species]:
                        self.logs[species][grade] = {rng: {sub: {'gross': [], 'mean': 0} for sub in subs} for rng in LOG_LENGTHS}
                        self.logs[species][grade]['totals_by_grade'] = {sub: {'gross': [], 'mean': 0} for sub in subs}
                    for rng in plot.logs[species][grade]:
                        if rng != 'display':
                            for sub in subs:
                                self.logs[species][grade][rng][sub]['gross'].append(plot.logs[species][grade][rng][sub])
            for species in self.logs:
                for grade in self.logs[species]:
                    for rng in self.logs[species][grade]:
                        for sub in subs:
                            gross = self.logs[species][grade][rng][sub]['gross']
                            if len(gross) < self.plot_count:
                                gross += ([0] * (self.plot_count - len(gross)))
                            self.logs[species][grade][rng][sub]['mean'] = mean(gross)

    def _get_stats(self, data):
        """Runs the statistical calculations on a set of the stand conditions data, returns an updated sub dict, used internally"""
        m = mean(data)
        if len(data) >= 2:
            std = stdev(data)
            ste = std / math.sqrt(self.plot_count)
            low_avg_high = [max(round(m - ste, 1), 0), m, m + ste]
            d = {'mean': m,
                 'variance': variance(data),
                 'stdev': std,
                 'stderr': ste,
                 'stderr_pct': (ste / m) * 100,
                 'low_avg_high': low_avg_high}
        else:
            d = {'mean': m,
                 'variance': 'Not enough data',
                 'stdev': 'Not enough data',
                 'stderr': 'Not enough data',
                 'stderr_pct': 'Not enough data',
                 'low_avg_high': 'Not enough data'}
        return d

    # def _update_table_data(self):
    #     """Converts stand data to plot/tree inventory data table layout, used internally"""
    #     heads = ['Stand', 'Plot Number', 'Tree Number', 'Species', 'DBH', 'Height',
    #              'Stump Height', 'Log 1 Length', 'Log 1 Grade', 'Log 1 Defect', 'Between Logs Feet']
    #     master = []
    #     max_logs = []
    #     for i, plot in enumerate(self.plots):
    #         for j, tree in enumerate(plot.trees):
    #             temp = [self.name, i + 1, j + 1]
    #             for key in ['species', 'dbh', 'height']:
    #                 temp.append(tree[key])
    #             len_logs = len(tree.logs)
    #             max_logs.append(len_logs)
    #             for k, lnum in enumerate(tree.logs):
    #                 log = tree.logs[lnum]
    #                 if lnum == 1:
    #                     temp.append(log.stem_height - log.length - 1)
    #                 for lkey in ['length', 'grade', 'defect']:
    #                     temp.append(log[lkey])
    #                 if k < len(tree.logs) - 1:
    #                     between = tree.logs[lnum+1].stem_height - log.stem_height - tree.logs[lnum+1].length - 1
    #                     if between < 0:
    #                         temp.append(0)
    #                     else:
    #                         temp.append(between)
    #             master.append(temp)
    #
    #     heads += add_logs_to_table_heads(max(max_logs))
    #     len_heads = len(heads)
    #     for i in master:
    #         len_i = len(i)
    #         if len_i < len_heads:
    #             i += ['' for j in range(len_heads - len_i)]
    #     master.insert(0, heads)
    #     return master
    #
    # def _update_summary_stand(self):
    #     """Updates the current stand conditions list of stand.summary_stand, used internally"""
    #     heads = ['SPECIES'] + [head[1] for head in SORTED_HEADS]
    #     body_data = []
    #     for key in self.species:
    #         if key == 'totals_all':
    #             show = 'TOTALS'
    #         else:
    #             show = key
    #         temp = [str(show)] + [format_comma(self.species[key][i[0]]) for i in SORTED_HEADS]
    #         body_data.append(temp)
    #     body_data.append(body_data.pop(0))
    #     body_data.insert(0, heads)
    #     return body_data
    #
    # def _update_summary_logs(self):
    #     """Updates the stand logs summary dict, data-tables are broken down by metric type --> species, used internally.
    #     Example: self.summary_logs['BOARD FEET PER ACRE']['DF'] --> data table"""
    #     table_data = {}
    #     tables = [['bf_ac', 'BOARD FEET PER ACRE'], ['cf_ac', 'CUBIC FEET PER ACRE'], ['lpa', 'LOGS PER ACRE']]
    #     for table in tables:
    #         metric_key = table[0]
    #         key = table[1]
    #         table_data[key] = {}
    #         for species in self.logs:
    #             if species == 'totals_all':
    #                 show = 'TOTALS'
    #             else:
    #                 show = ALL_SPECIES_NAMES[species]
    #
    #             table_data[key][show] = [['LOG GRADES'] + [rng.upper() for rng in LOG_LENGTHS] + ['TOTALS']]
    #
    #             grade_sort = []
    #             for grade in self.logs[species]:
    #                 values = [self.logs[species][grade][rng][metric_key]['mean'] for rng in self.logs[species][grade]]
    #                 if sum(values) > 0:
    #                     if grade == 'totals_by_length':
    #                         col_text = 'TOTALS'
    #                     else:
    #                         col_text = grade
    #                     grade_sort.append([col_text] + [format_comma(z) for z in values])
    #             grade_sort = sorted(grade_sort, key=lambda x: GRADE_SORT[x[0]])
    #             for g in grade_sort:
    #                 table_data[key][show].append(g)
    #
    #         table_data[key] = reorder_dict(table_data[key])
    #     return table_data
    #
    # def _update_summary_stats(self):
    #     """Updates the stand statistics dict, stats-tables are broken down by species, used internally.
    #     Example: self.summary_stats['DF'] --> stats-table"""
    #     tables = {}
    #     for spp in self.species_stats:
    #         if spp == 'totals_all':
    #             show = 'TOTALS'
    #         else:
    #             show = ALL_SPECIES_NAMES[spp]
    #
    #         tables[show] = [['METRIC'] + [head.upper() for head in self.species_stats[spp]['tpa'] if head != 'low_avg_high'] + ['LOW',
    #                                                                                                                             'AVERAGE',
    #                                                                                                                             'HIGH']]
    #         for key in self.species_stats[spp]:
    #             temp = [key.upper()]
    #             not_enough_data = False
    #             for sub in self.species_stats[spp][key]:
    #                 x = self.species_stats[spp][key][sub]
    #                 if not_enough_data:
    #                     if x == 'Not enough data':
    #                         if sub == 'low_avg_high':
    #                             for i in range(3):
    #                                 temp.append('-')
    #                         else:
    #                             temp.append('-')
    #                 else:
    #                     if x == 'Not enough data':
    #                         temp.append(x)
    #                         not_enough_data = True
    #                     else:
    #                         if sub == 'low_avg_high':
    #                             for i in x:
    #                                 temp.append(format_comma(i))
    #                         elif sub == 'stderr_pct':
    #                             temp.append(format_pct(x))
    #                         else:
    #                             temp.append(format_comma(x))
    #             tables[show].append(temp)
    #     return reorder_dict(tables)



    # def _compile_report_text(self):
    #     """Compiles the console-formatted report of all stand data and stats, used internally"""
    #     n = '\n' * 4
    #     console_text = f'{print_stand_species(self.summary_stand)}{n}'
    #     console_text += f'{print_stand_logs(self.summary_logs)}{n}'
    #     console_text += f'{print_stand_stats(self.summary_stats)}'
    #     return console_text
    #
    # def _compile_pdf_report(self):
    #     pdf = PDF()
    #     pdf.alias_nb_pages()
    #     pdf.add_page()
    #     pdf.compile_stand_report(self)
    #     return pdf



if __name__ == '__main__':
    pass




